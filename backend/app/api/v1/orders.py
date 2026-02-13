"""
订单管理 API
"""

from typing import List, Dict, Any
from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from datetime import date, datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from urllib.parse import quote

from app.core.database import get_db
from app.core.exceptions import BusinessException, NotFoundException
from app.api.deps import get_current_user, check_permission
from app.models.user import User
from app.models.order import OrderHeader
from app.models.store import Store
from app.schemas.common import Response, success
from app.schemas.order import OrderCreate
from app.services.data_scope_service import assert_store_access
from app.services.audit import create_audit_log
from app.services.order_service import get_order_list, get_order_export_rows

router = APIRouter()


@router.get(
    "",
    response_model=Response[Dict[str, Any]],
    summary="获取订单列表",
    description="获取订单列表"
)
async def list_orders(
    store_id: int = Query(None, description="门店ID"),
    channel: str = Query(None, description="渠道"),
    order_no: str = Query(None, description="订单号"),
    start_date: date = Query(None, description="开始日期"),
    end_date: date = Query(None, description="结束日期"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页大小"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取订单列表"""
    result = await get_order_list(
        db=db,
        current_user=current_user,
        store_id=store_id,
        channel=channel,
        order_no=order_no,
        start_date=start_date,
        end_date=end_date,
        page=page,
        page_size=page_size,
    )
    
    return success(
        data={
            "items": result["items"],
            "total": result["total"],
            "page": result["page"],
            "page_size": result["page_size"]
        }
    )


@router.get(
    "/export",
    summary="导出订单",
    description="导出订单列表为Excel文件"
)
async def export_orders(
    request: Request,
    store_id: int = Query(None, description="门店ID"),
    channel: str = Query(None, description="渠道"),
    order_no: str = Query(None, description="订单号"),
    start_date: date = Query(None, description="开始日期"),
    end_date: date = Query(None, description="结束日期"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """导出订单列表"""
    await check_permission(current_user, "order:export", db)
    rows = await get_order_export_rows(
        db=db,
        current_user=current_user,
        store_id=store_id,
        channel=channel,
        order_no=order_no,
        start_date=start_date,
        end_date=end_date,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单列表"

    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["序号", "订单号", "门店", "渠道", "金额", "订单时间", "备注", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for idx, row in enumerate(rows, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=row.OrderHeader.order_no)
        ws.cell(row=idx + 1, column=3, value=row.store_name or "未知门店")
        ws.cell(row=idx + 1, column=4, value=row.OrderHeader.channel or "未知")
        ws.cell(row=idx + 1, column=5, value=float(row.OrderHeader.net_amount or 0))
        ws.cell(
            row=idx + 1,
            column=6,
            value=row.OrderHeader.order_time.strftime("%Y-%m-%d %H:%M:%S") if row.OrderHeader.order_time else ""
        )
        ws.cell(row=idx + 1, column=7, value=row.OrderHeader.remark or "")
        ws.cell(row=idx + 1, column=8, value=row.OrderHeader.status or "completed")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 12

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    await create_audit_log(
        db=db,
        user=current_user,
        action="EXPORT_ORDERS",
        resource="order",
        resource_id=None,
        detail={
            "store_id": store_id,
            "channel": channel,
            "order_no": order_no,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "count": len(rows)
        },
        request=request,
        status_code=200
    )
    await db.commit()

    filename = f"订单列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=orders.xlsx; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.get(
    "/{order_id}",
    response_model=Response[dict],
    summary="获取订单详情",
    description="获取订单详情"
)
async def get_order(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取订单详情"""
    result = await db.execute(select(OrderHeader).where(OrderHeader.id == order_id))
    order = result.scalar_one_or_none()
    
    if not order:
        raise NotFoundException("订单不存在")
    
    # 数据权限校验：检查是否有权访问该门店的订单
    await assert_store_access(db, current_user, order.store_id)
    
    return success(
        data={
            "id": order.id,
            "order_no": order.order_no,
            "total_amount": float(order.total_amount),
            "status": order.status,
            "created_at": order.created_at.isoformat()
        }
    )


@router.post(
    "",
    response_model=Response[dict],
    summary="创建订单",
    description="创建新的订单"
)
async def create_order(
    data: OrderCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """创建订单"""
    # 数据权限校验：检查是否有权访问该门店
    await assert_store_access(db, current_user, data.store_id)
    
    # 验证门店存在
    store_result = await db.execute(select(Store).where(Store.id == data.store_id))
    if not store_result.scalar_one_or_none():
        raise NotFoundException("门店不存在")
    
    # 检查订单号是否已存在
    existing_order = await db.execute(
        select(OrderHeader).where(OrderHeader.order_no == data.order_no)
    )
    if existing_order.scalar_one_or_none():
        raise BusinessException("订单号已存在")
    
    # 创建订单
    order = OrderHeader(
        store_id=data.store_id,
        channel=data.channel,
        order_no=data.order_no,
        net_amount=data.net_amount,
        total_amount=data.net_amount,  # 简化处理，实际可能需要加上折扣等
        order_time=data.order_time,
        remark=data.remark,
        status="completed"
    )
    
    db.add(order)
    await db.commit()
    await db.refresh(order)
    
    # 记录审计日志
    await create_audit_log(
        db=db,
        user=current_user,
        action="CREATE_ORDER",
        resource="order",
        resource_id=str(order.id),
        detail={
            "store_id": data.store_id,
            "order_no": data.order_no,
            "channel": data.channel,
            "net_amount": float(data.net_amount)
        },
        request=request,
        status_code=201
    )
    await db.commit()
    
    return success(
        data={
            "id": order.id,
            "order_no": order.order_no,
            "store_id": order.store_id,
            "channel": order.channel,
            "amount": float(order.net_amount),
            "order_time": order.order_time.isoformat(),
            "remark": order.remark,
            "created_at": order.created_at.isoformat()
        }
    )