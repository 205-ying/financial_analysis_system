"""
费用记录管理 API
"""

from typing import List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, status, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from datetime import date, datetime
from pydantic import BaseModel, Field
from decimal import Decimal
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from urllib.parse import quote

from app.core.database import get_db
from app.api.deps import get_current_user, check_permission
from app.models.user import User
from app.models.expense import ExpenseRecord, ExpenseType
from app.models.store import Store
from app.schemas.common import Response, success
from app.services.audit import create_audit_log
from app.services.data_scope_service import filter_stores_by_access, assert_store_access

router = APIRouter()


# Schemas
class ExpenseRecordCreate(BaseModel):
    """创建费用记录请求"""
    store_id: int = Field(..., description="门店ID")
    expense_type_id: int = Field(..., description="费用类型ID")
    biz_date: date = Field(..., description="业务日期")
    amount: Decimal = Field(..., description="金额")
    remark: str = Field("", description="备注")


class ExpenseRecordUpdate(BaseModel):
    """更新费用记录请求"""
    store_id: int = Field(None, description="门店ID")
    expense_type_id: int = Field(None, description="费用类型ID")
    biz_date: date = Field(None, description="业务日期")
    amount: Decimal = Field(None, description="金额")
    remark: str = Field(None, description="备注")


@router.get(
    "",
    response_model=Response[Dict[str, Any]],
    summary="获取费用记录列表",
    description="获取费用记录列表"
)
async def list_expense_records(
    store_id: int = Query(None, description="门店ID"),
    expense_type_id: int = Query(None, description="费用类型ID"),
    start_date: date = Query(None, description="开始日期"),
    end_date: date = Query(None, description="结束日期"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页大小"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取费用记录列表"""
    # 数据权限过滤：获取可访问的门店ID列表
    accessible_store_ids = await filter_stores_by_access(db, current_user, store_id)
    
    # 构建查询，联查门店和费用类型
    query = select(
        ExpenseRecord,
        Store.name.label('store_name'),
        ExpenseType.name.label('expense_type_name')
    ).join(
        Store, ExpenseRecord.store_id == Store.id, isouter=True
    ).join(
        ExpenseType, ExpenseRecord.expense_type_id == ExpenseType.id, isouter=True
    )
    
    # 应用数据权限过滤
    conditions = []
    if accessible_store_ids is not None:
        conditions.append(ExpenseRecord.store_id.in_(accessible_store_ids))
    
    # 应用筛选条件
    if expense_type_id:
        conditions.append(ExpenseRecord.expense_type_id == expense_type_id)
    if start_date:
        conditions.append(ExpenseRecord.biz_date >= start_date)
    if end_date:
        conditions.append(ExpenseRecord.biz_date <= end_date)
    
    if conditions:
        query = query.where(and_(*conditions))
    
    # 获取总数
    count_query = select(func.count()).select_from(ExpenseRecord)
    if conditions:
        count_query = count_query.where(and_(*conditions))
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0
    
    # 应用分页
    offset = (page - 1) * page_size
    query = query.order_by(ExpenseRecord.biz_date.desc()).offset(offset).limit(page_size)
    
    result = await db.execute(query)
    rows = result.all()
    
    items = [{
        "id": row.ExpenseRecord.id,
        "store_id": row.ExpenseRecord.store_id,
        "store_name": row.store_name or '未知门店',
        "expense_type_id": row.ExpenseRecord.expense_type_id,
        "expense_type_name": row.expense_type_name or '未知类型',
        "expense_date": row.ExpenseRecord.biz_date.isoformat(),
        "amount": float(row.ExpenseRecord.amount or 0),
        "remark": row.ExpenseRecord.remark or '',
        "created_at": row.ExpenseRecord.created_at.isoformat() if row.ExpenseRecord.created_at else ''
    } for row in rows]
    
    return success(
        data={
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size
        }
    )


@router.get(
    "/export",
    summary="导出费用记录",
    description="导出费用记录列表为Excel文件"
)
async def export_expense_records(
    request: Request,
    store_id: int = Query(None, description="门店ID"),
    expense_type_id: int = Query(None, description="费用类型ID"),
    start_date: date = Query(None, description="开始日期"),
    end_date: date = Query(None, description="结束日期"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """导出费用记录列表"""
    # 权限检查
    await check_permission(current_user, "expense:export", db)
    
    # 数据权限过滤
    accessible_store_ids = await filter_stores_by_access(db, current_user, store_id)
    
    # 构建查询
    query = select(
        ExpenseRecord,
        Store.name.label('store_name'),
        ExpenseType.name.label('expense_type_name'),
        ExpenseType.category.label('expense_type_category')
    ).join(
        Store, ExpenseRecord.store_id == Store.id, isouter=True
    ).join(
        ExpenseType, ExpenseRecord.expense_type_id == ExpenseType.id, isouter=True
    )
    
    # 应用数据权限过滤
    conditions = []
    if accessible_store_ids is not None:
        conditions.append(ExpenseRecord.store_id.in_(accessible_store_ids))
    
    # 应用筛选条件
    if expense_type_id:
        conditions.append(ExpenseRecord.expense_type_id == expense_type_id)
    if start_date:
        conditions.append(ExpenseRecord.biz_date >= start_date)
    if end_date:
        conditions.append(ExpenseRecord.biz_date <= end_date)
    
    if conditions:
        query = query.where(and_(*conditions))
    
    # 排序并限制数量（导出最多10000条）
    query = query.order_by(ExpenseRecord.biz_date.desc()).limit(10000)
    
    result = await db.execute(query)
    rows = result.all()
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "费用记录"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ["序号", "门店", "费用类型", "费用分类", "金额", "费用日期", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据
    for idx, row in enumerate(rows, 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=row.store_name or '未知门店')
        ws.cell(row=idx+1, column=3, value=row.expense_type_name or '未知类型')
        ws.cell(row=idx+1, column=4, value=row.expense_type_category or '未分类')
        ws.cell(row=idx+1, column=5, value=float(row.ExpenseRecord.amount or 0))
        ws.cell(row=idx+1, column=6, value=row.ExpenseRecord.biz_date.strftime('%Y-%m-%d') if row.ExpenseRecord.biz_date else '')
        ws.cell(row=idx+1, column=7, value=row.ExpenseRecord.remark or '')
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    
    # 保存到内存
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 记录审计日志
    await create_audit_log(
        db=db,
        user=current_user,
        action="EXPORT_EXPENSES",
        resource="expense",
        resource_id=None,
        detail={
            "store_id": store_id,
            "expense_type_id": expense_type_id,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "count": len(rows)
        },
        request=request,
        status_code=200
    )
    await db.commit()
    
    # 返回Excel文件
    filename = f"费用记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=expenses.xlsx; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.get(
    "/{record_id}",
    response_model=Response[dict],
    summary="获取费用记录详情",
    description="获取费用记录详情"
)
async def get_expense_record(
    record_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """获取费用记录详情"""
    result = await db.execute(select(ExpenseRecord).where(ExpenseRecord.id == record_id))
    record = result.scalar_one_or_none()
    
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="费用记录不存在"
        )
    
    # 数据权限校验：检查是否有权访问该门店的数据
    await assert_store_access(db, current_user, record.store_id)
    
    return success(
        data={
            "id": record.id,
            "store_id": record.store_id,
            "expense_type_id": record.expense_type_id,
            "business_date": record.business_date.isoformat(),
            "amount": float(record.amount),
            "description": record.description,
            "vendor": record.vendor,
            "payment_method": record.payment_method,
            "created_at": record.created_at.isoformat(),
            "updated_at": record.updated_at.isoformat()
        }
    )


@router.post(
    "",
    response_model=Response[dict],
    summary="创建费用记录",
    description="创建新的费用记录"
)
async def create_expense_record(
    data: ExpenseRecordCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """创建费用记录"""
    # 数据权限校验：检查是否有权访问该门店
    await assert_store_access(db, current_user, data.store_id)
    
    # 验证门店存在
    store_result = await db.execute(select(Store).where(Store.id == data.store_id))
    if not store_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="门店不存在"
        )
    
    # 验证费用类型存在
    type_result = await db.execute(select(ExpenseType).where(ExpenseType.id == data.expense_type_id))
    if not type_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="费用类型不存在"
        )
    
    # 创建费用记录
    record = ExpenseRecord(
        store_id=data.store_id,
        expense_type_id=data.expense_type_id,
        biz_date=data.biz_date,
        amount=data.amount,
        remark=data.remark
    )
    
    db.add(record)
    await db.commit()
    await db.refresh(record)
    
    # 记录审计日志
    await create_audit_log(
        db=db,
        user=current_user,
        action="CREATE_EXPENSE",
        resource="expense",
        resource_id=str(record.id),
        detail={
            "store_id": data.store_id,
            "expense_type_id": data.expense_type_id,
            "biz_date": data.biz_date.isoformat(),
            "amount": float(data.amount)
        },
        request=request,
        status_code=201
    )
    await db.commit()
    
    return success(
        data={
            "id": record.id,
            "store_id": record.store_id,
            "expense_type_id": record.expense_type_id,
            "biz_date": record.biz_date.isoformat(),
            "amount": float(record.amount),
            "remark": record.remark,
            "created_at": record.created_at.isoformat()
        }
    )


@router.put(
    "/{record_id}",
    response_model=Response[dict],
    summary="更新费用记录",
    description="更新费用记录信息"
)
async def update_expense_record(
    record_id: int,
    data: ExpenseRecordUpdate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """更新费用记录"""
    # 查询记录
    result = await db.execute(select(ExpenseRecord).where(ExpenseRecord.id == record_id))
    record = result.scalar_one_or_none()
    
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="费用记录不存在"
        )
    
    # 数据权限校验：检查是否有权访问原门店
    await assert_store_access(db, current_user, record.store_id)
    
    # 如果要更改门店，校验新门店权限
    if data.store_id is not None and data.store_id != record.store_id:
        await assert_store_access(db, current_user, data.store_id)
    
    # 保存旧值用于审计
    old_values = {
        "store_id": record.store_id,
        "expense_type_id": record.expense_type_id,
        "biz_date": record.biz_date.isoformat(),
        "amount": float(record.amount),
        "remark": record.remark
    }
    
    # 更新字段
    if data.store_id is not None:
        # 验证门店存在
        store_result = await db.execute(select(Store).where(Store.id == data.store_id))
        if not store_result.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="门店不存在"
            )
        record.store_id = data.store_id
    
    if data.expense_type_id is not None:
        # 验证费用类型存在
        type_result = await db.execute(select(ExpenseType).where(ExpenseType.id == data.expense_type_id))
        if not type_result.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="费用类型不存在"
            )
        record.expense_type_id = data.expense_type_id
    
    if data.biz_date is not None:
        record.biz_date = data.biz_date
    
    if data.amount is not None:
        record.amount = data.amount
    
    if data.remark is not None:
        record.remark = data.remark
    
    await db.commit()
    await db.refresh(record)
    
    # 记录审计日志
    new_values = {
        "store_id": record.store_id,
        "expense_type_id": record.expense_type_id,
        "biz_date": record.biz_date.isoformat(),
        "amount": float(record.amount),
        "remark": record.remark
    }
    
    await create_audit_log(
        db=db,
        user=current_user,
        action="UPDATE_EXPENSE",
        resource="expense",
        resource_id=str(record.id),
        detail={
            "old": old_values,
            "new": new_values
        },
        request=request,
        status_code=200
    )
    await db.commit()
    
    return success(
        data={
            "id": record.id,
            "store_id": record.store_id,
            "expense_type_id": record.expense_type_id,
            "biz_date": record.biz_date.isoformat(),
            "amount": float(record.amount),
            "remark": record.remark,
            "updated_at": record.updated_at.isoformat()
        }
    )


@router.delete(
    "/{record_id}",
    response_model=Response[dict],
    summary="删除费用记录",
    description="删除费用记录"
)
async def delete_expense_record(
    record_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """删除费用记录"""
    # 查询记录
    result = await db.execute(select(ExpenseRecord).where(ExpenseRecord.id == record_id))
    record = result.scalar_one_or_none()
    
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="费用记录不存在"
        )
    
    # 数据权限校验：检查是否有权访问该门店
    await assert_store_access(db, current_user, record.store_id)
    
    # 保存记录信息用于审计
    record_info = {
        "store_id": record.store_id,
        "expense_type_id": record.expense_type_id,
        "biz_date": record.biz_date.isoformat(),
        "amount": float(record.amount),
        "remark": record.remark
    }
    
    # 删除记录
    await db.delete(record)
    await db.commit()
    
    # 记录审计日志
    await create_audit_log(
        db=db,
        user=current_user,
        action="DELETE_EXPENSE",
        resource="expense",
        resource_id=str(record_id),
        detail=record_info,
        request=request,
        status_code=200
    )
    await db.commit()
    
    return success(data={"message": "费用记录已删除"})

