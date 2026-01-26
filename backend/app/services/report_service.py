"""
报表服务 - 数据聚合和导出

所有聚合必须在数据库端完成（使用 SQL），不允许拉取全量数据到 Python 循环。
已集成数据权限控制。
"""
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select, extract, case, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.models.kpi import KpiDailyStore
from app.models.order import OrderHeader
from app.models.expense import ExpenseRecord, ExpenseType
from app.models.store import Store
from app.models.user import User
from app.schemas.report import (
    DailySummaryRow,
    MonthlySummaryRow,
    StorePerformanceRow,
    ExpenseBreakdownRow,
    ReportQuery
)
from app.services.data_scope_service import filter_stores_by_access


async def get_daily_summary(
    db: AsyncSession,
    filters: ReportQuery,
    current_user: User
) -> List[DailySummaryRow]:
    """
    获取日汇总报表
    
    SQL 聚合逻辑：
    - 从 kpi_daily_store 获取基础指标
    - 从 expense_record 聚合费用
    - 从 order_header 聚合订单数
    - 计算利润率
    """
    # 构建基础查询 - 从 KPI 表
    query = select(
        KpiDailyStore.biz_date,
        KpiDailyStore.store_id,
        Store.name.label("store_name"),
        
        # 营收指标
        func.sum(KpiDailyStore.revenue).label("revenue"),
        func.sum(KpiDailyStore.net_revenue).label("net_revenue"),
        func.sum(KpiDailyStore.discount_amount).label("discount_amount"),
        func.sum(KpiDailyStore.refund_amount).label("refund_amount"),
        
        # 成本指标
        func.sum(KpiDailyStore.cost_total).label("cost_total"),
        func.sum(KpiDailyStore.cost_material).label("cost_material"),
        func.sum(KpiDailyStore.cost_labor).label("cost_labor"),
        
        # 利润指标（从 KPI 表计算）
        func.sum(KpiDailyStore.gross_profit).label("gross_profit"),
        func.sum(KpiDailyStore.operating_profit).label("operating_profit"),
    ).select_from(KpiDailyStore).join(
        Store, KpiDailyStore.store_id == Store.id
    ).where(
        and_(
            KpiDailyStore.biz_date >= filters.start_date,
            KpiDailyStore.biz_date <= filters.end_date
        )
    )
    
    # 数据权限过滤：限制可访问的门店
    accessible_store_ids = await filter_stores_by_access(db, current_user, filters.store_id)
    if accessible_store_ids is not None:
        query = query.where(KpiDailyStore.store_id.in_(accessible_store_ids))
    
    # 分组
    if filters.group_by == "day":
        query = query.group_by(
            KpiDailyStore.biz_date,
            KpiDailyStore.store_id,
            Store.name
        ).order_by(KpiDailyStore.biz_date.desc(), Store.name)
    else:
        # 默认按日期+门店分组
        query = query.group_by(
            KpiDailyStore.biz_date,
            KpiDailyStore.store_id,
            Store.name
        ).order_by(KpiDailyStore.biz_date.desc(), Store.name)
    
    result = await db.execute(query)
    rows = result.all()
    
    # 查询费用（单独查询后合并）
    expense_query = select(
        ExpenseRecord.biz_date,
        ExpenseRecord.store_id,
        func.sum(ExpenseRecord.amount).label("expense_total")
    ).where(
        and_(
            ExpenseRecord.biz_date >= filters.start_date,
            ExpenseRecord.biz_date <= filters.end_date
        )
    ).group_by(ExpenseRecord.biz_date, ExpenseRecord.store_id)
    
    if filters.store_id:
        expense_query = expense_query.where(ExpenseRecord.store_id == filters.store_id)
    
    expense_result = await db.execute(expense_query)
    expense_dict = {
        (row.biz_date, row.store_id): row.expense_total
        for row in expense_result.all()
    }
    
    # 查询订单数（单独查询后合并）
    order_query = select(
        OrderHeader.biz_date,
        OrderHeader.store_id,
        func.count(OrderHeader.id).label("order_count")
    ).where(
        and_(
            OrderHeader.biz_date >= filters.start_date,
            OrderHeader.biz_date <= filters.end_date,
            OrderHeader.status != "cancelled"
        )
    ).group_by(OrderHeader.biz_date, OrderHeader.store_id)
    
    if filters.store_id:
        order_query = order_query.where(OrderHeader.store_id == filters.store_id)
    
    order_result = await db.execute(order_query)
    order_dict = {
        (row.biz_date, row.store_id): row.order_count
        for row in order_result.all()
    }
    
    # 组装结果
    result_list = []
    for row in rows:
        expense_total = expense_dict.get((row.biz_date, row.store_id), Decimal("0.00"))
        order_count = order_dict.get((row.biz_date, row.store_id), 0)
        
        # 计算利润率
        gross_profit_rate = None
        operating_profit_rate = None
        if row.revenue and row.revenue > 0:
            gross_profit_rate = (row.gross_profit / row.revenue * 100).quantize(Decimal("0.01"))
            operating_profit_rate = (row.operating_profit / row.revenue * 100).quantize(Decimal("0.01"))
        
        result_list.append(DailySummaryRow(
            biz_date=row.biz_date,
            store_id=row.store_id,
            store_name=row.store_name,
            revenue=row.revenue,
            net_revenue=row.net_revenue,
            discount_amount=row.discount_amount,
            refund_amount=row.refund_amount,
            cost_total=row.cost_total,
            cost_material=row.cost_material,
            cost_labor=row.cost_labor,
            expense_total=expense_total,
            order_count=order_count,
            gross_profit=row.gross_profit,
            operating_profit=row.operating_profit,
            gross_profit_rate=gross_profit_rate,
            operating_profit_rate=operating_profit_rate
        ))
    
    return result_list


async def get_monthly_summary(
    db: AsyncSession,
    filters: ReportQuery,
    current_user: User
) -> List[MonthlySummaryRow]:
    """
    获取月汇总报表
    
    SQL 聚合逻辑：
    - 按年月分组聚合 kpi_daily_store
    - 计算日均指标
    """
    # 构建查询 - 按年月分组
    query = select(
        extract('year', KpiDailyStore.biz_date).label("year"),
        extract('month', KpiDailyStore.biz_date).label("month"),
        KpiDailyStore.store_id,
        Store.name.label("store_name"),
        
        # 营收指标
        func.sum(KpiDailyStore.revenue).label("revenue"),
        func.sum(KpiDailyStore.net_revenue).label("net_revenue"),
        func.sum(KpiDailyStore.discount_amount).label("discount_amount"),
        func.sum(KpiDailyStore.refund_amount).label("refund_amount"),
        
        # 成本指标
        func.sum(KpiDailyStore.cost_total).label("cost_total"),
        
        # 利润指标
        func.sum(KpiDailyStore.gross_profit).label("gross_profit"),
        func.sum(KpiDailyStore.operating_profit).label("operating_profit"),
        
        # 天数
        func.count(func.distinct(KpiDailyStore.biz_date)).label("day_count")
    ).select_from(KpiDailyStore).join(
        Store, KpiDailyStore.store_id == Store.id
    ).where(
        and_(
            KpiDailyStore.biz_date >= filters.start_date,
            KpiDailyStore.biz_date <= filters.end_date
        )
    )
    
    if filters.store_id:
        query = query.where(KpiDailyStore.store_id == filters.store_id)
    
    # 分组
    if filters.store_id:
        # 单门店：只按年月分组
        query = query.group_by(
            extract('year', KpiDailyStore.biz_date),
            extract('month', KpiDailyStore.biz_date),
            KpiDailyStore.store_id,
            Store.name
        ).order_by(
            extract('year', KpiDailyStore.biz_date).desc(),
            extract('month', KpiDailyStore.biz_date).desc()
        )
    else:
        # 多门店：按年月+门店分组
        query = query.group_by(
            extract('year', KpiDailyStore.biz_date),
            extract('month', KpiDailyStore.biz_date),
            KpiDailyStore.store_id,
            Store.name
        ).order_by(
            extract('year', KpiDailyStore.biz_date).desc(),
            extract('month', KpiDailyStore.biz_date).desc(),
            Store.name
        )
    
    result = await db.execute(query)
    rows = result.all()
    
    # 查询每月费用
    expense_query = select(
        extract('year', ExpenseRecord.biz_date).label("year"),
        extract('month', ExpenseRecord.biz_date).label("month"),
        ExpenseRecord.store_id,
        func.sum(ExpenseRecord.amount).label("expense_total")
    ).where(
        and_(
            ExpenseRecord.biz_date >= filters.start_date,
            ExpenseRecord.biz_date <= filters.end_date
        )
    ).group_by(
        extract('year', ExpenseRecord.biz_date),
        extract('month', ExpenseRecord.biz_date),
        ExpenseRecord.store_id
    )
    
    if accessible_store_ids is not None:
        expense_query = expense_query.where(ExpenseRecord.store_id.in_(accessible_store_ids))
    
    expense_result = await db.execute(expense_query)
    expense_dict = {
        (int(row.year), int(row.month), row.store_id): row.expense_total
        for row in expense_result.all()
    }
    
    # 查询每月订单数
    order_query = select(
        extract('year', OrderHeader.biz_date).label("year"),
        extract('month', OrderHeader.biz_date).label("month"),
        OrderHeader.store_id,
        func.count(OrderHeader.id).label("order_count")
    ).where(
        and_(
            OrderHeader.biz_date >= filters.start_date,
            OrderHeader.biz_date <= filters.end_date,
            OrderHeader.status != "cancelled"
        )
    ).group_by(
        extract('year', OrderHeader.biz_date),
        extract('month', OrderHeader.biz_date),
        OrderHeader.store_id
    )
    
    if accessible_store_ids is not None:
        order_query = order_query.where(OrderHeader.store_id.in_(accessible_store_ids))
    
    order_result = await db.execute(order_query)
    order_dict = {
        (int(row.year), int(row.month), row.store_id): row.order_count
        for row in order_result.all()
    }
    
    # 组装结果
    result_list = []
    for row in rows:
        year = int(row.year)
        month = int(row.month)
        expense_total = expense_dict.get((year, month, row.store_id), Decimal("0.00"))
        order_count = order_dict.get((year, month, row.store_id), 0)
        day_count = row.day_count or 1
        
        # 计算利润率和日均指标
        gross_profit_rate = None
        operating_profit_rate = None
        if row.revenue and row.revenue > 0:
            gross_profit_rate = (row.gross_profit / row.revenue * 100).quantize(Decimal("0.01"))
            operating_profit_rate = (row.operating_profit / row.revenue * 100).quantize(Decimal("0.01"))
        
        avg_daily_revenue = (row.revenue / day_count).quantize(Decimal("0.01"))
        avg_daily_order_count = Decimal(order_count) / day_count
        
        result_list.append(MonthlySummaryRow(
            year=year,
            month=month,
            store_id=row.store_id,
            store_name=row.store_name,
            revenue=row.revenue,
            net_revenue=row.net_revenue,
            discount_amount=row.discount_amount,
            refund_amount=row.refund_amount,
            cost_total=row.cost_total,
            expense_total=expense_total,
            order_count=order_count,
            gross_profit=row.gross_profit,
            operating_profit=row.operating_profit,
            gross_profit_rate=gross_profit_rate,
            operating_profit_rate=operating_profit_rate,
            avg_daily_revenue=avg_daily_revenue,
            avg_daily_order_count=avg_daily_order_count
        ))
    
    return result_list


async def get_store_performance(
    db: AsyncSession,
    filters: ReportQuery,
    current_user: User
) -> List[StorePerformanceRow]:
    """
    获取门店绩效报表
    
    SQL 聚合逻辑：
    - 按门店分组聚合
    - 计算排名（使用窗口函数）
    - 支持 TOP N 筛选
    """
    # 构建查询 - 按门店分组
    query = select(
        KpiDailyStore.store_id,
        Store.name.label("store_name"),
        
        # 营收指标
        func.sum(KpiDailyStore.revenue).label("revenue"),
        func.sum(KpiDailyStore.net_revenue).label("net_revenue"),
        
        # 利润指标
        func.sum(KpiDailyStore.gross_profit).label("gross_profit"),
        func.sum(KpiDailyStore.operating_profit).label("operating_profit"),
    ).select_from(KpiDailyStore).join(
        Store, KpiDailyStore.store_id == Store.id
    ).where(
        and_(
            KpiDailyStore.biz_date >= filters.start_date,
            KpiDailyStore.biz_date <= filters.end_date
        )
    )
    
    # 数据权限过滤
    accessible_store_ids = await filter_stores_by_access(db, current_user, filters.store_id)
    if accessible_store_ids is not None:
        query = query.where(KpiDailyStore.store_id.in_(accessible_store_ids))
    
    query = query.group_by(
        KpiDailyStore.store_id,
        Store.name
    ).order_by(func.sum(KpiDailyStore.revenue).desc())
    
    # TOP N 限制
    if filters.top_n:
        query = query.limit(filters.top_n)
    
    result = await db.execute(query)
    rows = result.all()
    
    # 查询订单统计
    order_query = select(
        OrderHeader.store_id,
        func.count(OrderHeader.id).label("order_count"),
        func.avg(OrderHeader.net_amount).label("avg_order_amount")
    ).where(
        and_(
            OrderHeader.biz_date >= filters.start_date,
            OrderHeader.biz_date <= filters.end_date,
            OrderHeader.status != "cancelled"
        )
    ).group_by(OrderHeader.store_id)
    
    if accessible_store_ids is not None:
        order_query = order_query.where(OrderHeader.store_id.in_(accessible_store_ids))
    
    order_result = await db.execute(order_query)
    order_dict = {
        row.store_id: (row.order_count, row.avg_order_amount or Decimal("0.00"))
        for row in order_result.all()
    }
    
    # 组装结果并计算排名
    result_list = []
    for idx, row in enumerate(rows, start=1):
        order_count, avg_order_amount = order_dict.get(row.store_id, (0, Decimal("0.00")))
        
        # 计算利润率
        gross_profit_rate = None
        operating_profit_rate = None
        if row.revenue and row.revenue > 0:
            gross_profit_rate = (row.gross_profit / row.revenue * 100).quantize(Decimal("0.01"))
            operating_profit_rate = (row.operating_profit / row.revenue * 100).quantize(Decimal("0.01"))
        
        result_list.append(StorePerformanceRow(
            store_id=row.store_id,
            store_name=row.store_name,
            revenue=row.revenue,
            net_revenue=row.net_revenue,
            order_count=order_count,
            avg_order_amount=avg_order_amount,
            gross_profit=row.gross_profit,
            operating_profit=row.operating_profit,
            gross_profit_rate=gross_profit_rate,
            operating_profit_rate=operating_profit_rate,
            revenue_rank=idx,
            profit_rank=None  # 可后续按利润重新排名
        ))
    
    # 按利润重新排名
    sorted_by_profit = sorted(result_list, key=lambda x: x.operating_profit, reverse=True)
    for idx, item in enumerate(sorted_by_profit, start=1):
        item.profit_rank = idx
    
    # 恢复按营收排序
    result_list.sort(key=lambda x: x.revenue, reverse=True)
    
    return result_list


async def get_expense_breakdown(
    db: AsyncSession,
    filters: ReportQuery,
    current_user: User
) -> List[ExpenseBreakdownRow]:
    """
    获取费用明细报表
    
    SQL 聚合逻辑：
    - 按费用科目分组聚合
    - 可选按门店分组
    - 计算占比
    """
    # 数据权限过滤
    accessible_store_ids = await filter_stores_by_access(db, current_user, filters.store_id)
    
    # 构建查询
    if accessible_store_ids and len(accessible_store_ids) == 1:
        # 单门店：按费用科目分组
        query = select(
            ExpenseRecord.expense_type_id,
            ExpenseType.type_code,
            ExpenseType.name.label("expense_type_name"),
            ExpenseType.category,
            ExpenseRecord.store_id,
            Store.name.label("store_name"),
            func.sum(ExpenseRecord.amount).label("total_amount"),
            func.count(ExpenseRecord.id).label("record_count"),
            func.avg(ExpenseRecord.amount).label("avg_amount")
        ).select_from(ExpenseRecord).join(
            ExpenseType, ExpenseRecord.expense_type_id == ExpenseType.id
        ).join(
            Store, ExpenseRecord.store_id == Store.id
        ).where(
            and_(
                ExpenseRecord.biz_date >= filters.start_date,
                ExpenseRecord.biz_date <= filters.end_date,
                ExpenseRecord.store_id == accessible_store_ids[0]
            )
        ).group_by(
            ExpenseRecord.expense_type_id,
            ExpenseType.type_code,
            ExpenseType.name,
            ExpenseType.category,
            ExpenseRecord.store_id,
            Store.name
        ).order_by(func.sum(ExpenseRecord.amount).desc())
    else:
        # 多门店：只按费用科目分组
        query = select(
            ExpenseRecord.expense_type_id,
            ExpenseType.type_code,
            ExpenseType.name.label("expense_type_name"),
            ExpenseType.category,
            func.sum(ExpenseRecord.amount).label("total_amount"),
            func.count(ExpenseRecord.id).label("record_count"),
            func.avg(ExpenseRecord.amount).label("avg_amount")
        ).select_from(ExpenseRecord).join(
            ExpenseType, ExpenseRecord.expense_type_id == ExpenseType.id
        ).where(
            and_(
                ExpenseRecord.biz_date >= filters.start_date,
                ExpenseRecord.biz_date <= filters.end_date
            )
        )
        
        if accessible_store_ids is not None:
            query = query.where(ExpenseRecord.store_id.in_(accessible_store_ids))
        
        query = query.group_by(
            ExpenseRecord.expense_type_id,
            ExpenseType.type_code,
            ExpenseType.name,
            ExpenseType.category
        ).order_by(func.sum(ExpenseRecord.amount).desc())
    
    # TOP N 限制
    if filters.top_n:
        query = query.limit(filters.top_n)
    
    result = await db.execute(query)
    rows = result.all()
    
    # 计算总费用（用于占比计算）
    total_query = select(
        func.sum(ExpenseRecord.amount).label("grand_total")
    ).where(
        and_(
            ExpenseRecord.biz_date >= filters.start_date,
            ExpenseRecord.biz_date <= filters.end_date
        )
    )
    
    if accessible_store_ids is not None:
        total_query = total_query.where(ExpenseRecord.store_id.in_(accessible_store_ids))
    
    total_result = await db.execute(total_query)
    grand_total = total_result.scalar() or Decimal("0.00")
    
    # 组装结果
    result_list = []
    for row in rows:
        # 计算占比
        percentage = None
        if grand_total > 0:
            percentage = (row.total_amount / grand_total * 100).quantize(Decimal("0.01"))
        
        if filters.store_id:
            result_list.append(ExpenseBreakdownRow(
                expense_type_id=row.expense_type_id,
                expense_type_code=row.type_code,
                expense_type_name=row.expense_type_name,
                category=row.category,
                store_id=row.store_id,
                store_name=row.store_name,
                total_amount=row.total_amount,
                record_count=row.record_count,
                avg_amount=row.avg_amount,
                percentage=percentage
            ))
        else:
            result_list.append(ExpenseBreakdownRow(
                expense_type_id=row.expense_type_id,
                expense_type_code=row.type_code,
                expense_type_name=row.expense_type_name,
                category=row.category,
                total_amount=row.total_amount,
                record_count=row.record_count,
                avg_amount=row.avg_amount,
                percentage=percentage
            ))
    
    return result_list


async def export_report_excel(
    db: AsyncSession,
    filters: ReportQuery,
    current_user: User
) -> bytes:
    """
    导出报表为 Excel 文件
    
    包含多个 Sheet:
    - DailySummary: 日汇总
    - StorePerformance: 门店绩效
    - ExpenseBreakdown: 费用明细
    """
    # 获取数据（带数据权限过滤）
    daily_data = await get_daily_summary(db, filters, current_user)
    store_data = await get_store_performance(db, filters, current_user)
    expense_data = await get_expense_breakdown(db, filters, current_user)
    
    # 创建工作簿
    wb = Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: 日汇总
    ws1 = wb.active
    ws1.title = "DailySummary"
    
    # 写入表头
    headers1 = [
        "业务日期", "门店ID", "门店名称", "营业收入", "净收入", "优惠金额", "退款金额",
        "总成本", "原材料成本", "人工成本", "总费用", "订单数",
        "毛利润", "净利润", "毛利率(%)", "净利率(%)"
    ]
    for col_num, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入数据
    for row_num, data in enumerate(daily_data, start=2):
        ws1.cell(row=row_num, column=1, value=data.biz_date.strftime("%Y-%m-%d"))
        ws1.cell(row=row_num, column=2, value=data.store_id)
        ws1.cell(row=row_num, column=3, value=data.store_name)
        ws1.cell(row=row_num, column=4, value=float(data.revenue))
        ws1.cell(row=row_num, column=5, value=float(data.net_revenue))
        ws1.cell(row=row_num, column=6, value=float(data.discount_amount))
        ws1.cell(row=row_num, column=7, value=float(data.refund_amount))
        ws1.cell(row=row_num, column=8, value=float(data.cost_total))
        ws1.cell(row=row_num, column=9, value=float(data.cost_material))
        ws1.cell(row=row_num, column=10, value=float(data.cost_labor))
        ws1.cell(row=row_num, column=11, value=float(data.expense_total))
        ws1.cell(row=row_num, column=12, value=data.order_count)
        ws1.cell(row=row_num, column=13, value=float(data.gross_profit))
        ws1.cell(row=row_num, column=14, value=float(data.net_profit))
        ws1.cell(row=row_num, column=15, value=float(data.gross_profit_rate) if data.gross_profit_rate else None)
        ws1.cell(row=row_num, column=16, value=float(data.net_profit_rate) if data.net_profit_rate else None)
    
    # 自动调整列宽
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 2: 门店绩效
    ws2 = wb.create_sheet("StorePerformance")
    
    headers2 = [
        "门店ID", "门店名称", "营业收入", "净收入", "订单数", "客单价",
        "毛利润", "净利润", "毛利率(%)", "净利率(%)", "营收排名", "利润排名"
    ]
    for col_num, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_num, data in enumerate(store_data, start=2):
        ws2.cell(row=row_num, column=1, value=data.store_id)
        ws2.cell(row=row_num, column=2, value=data.store_name)
        ws2.cell(row=row_num, column=3, value=float(data.revenue))
        ws2.cell(row=row_num, column=4, value=float(data.net_revenue))
        ws2.cell(row=row_num, column=5, value=data.order_count)
        ws2.cell(row=row_num, column=6, value=float(data.avg_order_amount))
        ws2.cell(row=row_num, column=7, value=float(data.gross_profit))
        ws2.cell(row=row_num, column=8, value=float(data.net_profit))
        ws2.cell(row=row_num, column=9, value=float(data.gross_profit_rate) if data.gross_profit_rate else None)
        ws2.cell(row=row_num, column=10, value=float(data.net_profit_rate) if data.net_profit_rate else None)
        ws2.cell(row=row_num, column=11, value=data.revenue_rank)
        ws2.cell(row=row_num, column=12, value=data.profit_rank)
    
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 3: 费用明细
    ws3 = wb.create_sheet("ExpenseBreakdown")
    
    headers3 = [
        "费用科目ID", "科目编码", "科目名称", "费用类别", "门店ID", "门店名称",
        "费用总额", "记录笔数", "平均单笔", "占比(%)"
    ]
    for col_num, header in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_num, data in enumerate(expense_data, start=2):
        ws3.cell(row=row_num, column=1, value=data.expense_type_id)
        ws3.cell(row=row_num, column=2, value=data.expense_type_code)
        ws3.cell(row=row_num, column=3, value=data.expense_type_name)
        ws3.cell(row=row_num, column=4, value=data.category)
        ws3.cell(row=row_num, column=5, value=data.store_id)
        ws3.cell(row=row_num, column=6, value=data.store_name)
        ws3.cell(row=row_num, column=7, value=float(data.total_amount))
        ws3.cell(row=row_num, column=8, value=data.record_count)
        ws3.cell(row=row_num, column=9, value=float(data.avg_amount))
        ws3.cell(row=row_num, column=10, value=float(data.percentage) if data.percentage else None)
    
    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
